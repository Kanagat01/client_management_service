import json
from typing import List, Dict

import requests
import openpyxl
from urllib.parse import quote
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from rest_framework import status
from rest_framework.views import exception_handler
from rest_framework.exceptions import ValidationError
from rest_framework.response import Response
from backend.settings import WAPPI_PROFILE_ID, WAPPI_TOKEN


def error_with_text(text):
    return Response({'status': 'error', 'message': text}, status=status.HTTP_400_BAD_REQUEST)


def success_with_text(text):
    return Response({'status': 'success', 'message': text}, status=status.HTTP_200_OK)


def custom_exception_handler(exc, context):
    response = exception_handler(exc, context)

    if isinstance(exc, ValidationError):
        return error_with_text(exc.detail)

    return response


def export_to_excel(request, data: List[Dict], columns: Dict[str, str], title="Данные", filename="data_export") -> HttpResponse:
    """
    Экспорт данных в Excel.

    :param request: запрос Django (для совместимости с представлениями)
    :param data: список данных для экспорта (каждый элемент — это словарь)
    :param columns: словарь с названий столбцов и переводов
    :param title: название листа 
    :param filename: название файла
    :return: HttpResponse с файлом Excel
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title

    for col_num, column_title in enumerate(columns.values(), 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = column_title

    for row_num, item in enumerate(data, 2):
        for col_num, column in enumerate(columns.keys(), 1):
            col_letter = get_column_letter(col_num)
            field_value = item.get(column, '')
            ws[f'{col_letter}{row_num}'] = field_value

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    encoded_filename = quote(f"{filename}.xlsx")
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{encoded_filename}"
    wb.save(response)
    return response


def send_whatsapp_sms(phone: str, message: str):
    """
    Отправка сообщения в WhatsApp.

    :param phone: номер телефона
    :param message: текст сообщения
    """
    data = json.dumps({"body": message, "recipient": phone.replace("+", "")})
    response = requests.post(
        f"https://wappi.pro/api/sync/message/send?profile_id={WAPPI_PROFILE_ID}",
        headers={"Authorization": WAPPI_TOKEN},
        data=data
    )
    return response.json()
